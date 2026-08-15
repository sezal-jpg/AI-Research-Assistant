from pathlib import Path
from openpyxl import load_workbook
from langchain_core.documents import Document
from app.core.logger import logger
import pandas as pd

class ExcelLoader:
    def load(self,file_path:Path):
        logger.info(f'Loading Excel file: {file_path.name}')
        docs=[]
        suffix=file_path.suffix.lower()
        
        if suffix=='.xlsx':
            workbook=load_workbook(file_path,datat_only=True)
            for sheet in workbook.worksheets:
                rows=[]
                
                for row in sheet.iter_rows(values_only=True):
                    values=[str(value) for value in row if value is not None]
                    
                    if values:
                         rows.append("|".join(values))
                sheet_text="\n".join(rows)
                
                if sheet_text.strip():
                    docs.append(Document(page_content=sheet_text,metadata={'source_file':file_path.name,"source_type":'xlsx','sheet':sheet.title}))
                    
        elif suffix=='.xls':
            excel_file=pd.ExcelFile(file_path)
            for sheet_name in excel_file.sheet_names:
                dataframe=pd.read_excel(file_path,sheet_name=sheet_name)
                
                sheet_text=dataframe.to_string(index=False)
                
                if sheet_text.strip():
                    docs.append(Document(page_content=sheet_text,metadata={'source_file': file_path.name,'source_type':'xls','sheet':sheet_name,}))
                    
        logger.info(f'Loaded {len(docs)} Excel sheets')
        return docs
    
excel_loader=ExcelLoader()                
                
                                 